import io
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload

from app.database import get_db
from app.models.seat_allocation import SeatAllocation
from app.models.exam_session import ExamSession
from app.models.exam import Exam
from app.models.student import Student
from app.models.user import User
from app.schemas.seat_allocation import (
    GenerateSeatingRequest,
    GenerateSessionSeatingRequest,
    SeatingResponse,
    SeatAllocationDetail,
)
from app.services.seat_allocation import generate_seating, generate_session_seating
from app.auth import get_current_user

router = APIRouter()


# ─── Session-based seating generation (primary path) ─────────────────────────

@router.post("/generate-seating/session", status_code=status.HTTP_201_CREATED)
def generate_session_seating_route(
    payload: GenerateSessionSeatingRequest,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """
    Generate seating for a full exam session.

    Accepts multiple (exam_id, student_ids) pairs — one per department batch.
    Students from all batches are mixed across the selected halls.
    Guarantees no physical seat is double-booked within the session.
    """
    batches = [
        {"exam_id": b.exam_id, "student_ids": b.student_ids}
        for b in payload.batches
    ]
    try:
        allocations = generate_session_seating(
            session_id=payload.session_id,
            batches=batches,
            hall_ids=payload.hall_ids,
            db=db,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    halls_used = len({a.hall_id for a in allocations})
    return {
        "message": f"Seating generated. {len(allocations)} students allocated.",
        "total_allocated": len(allocations),
        "halls_used": halls_used,
        "session_id": payload.session_id,
    }


# ─── Legacy single-exam generation (backward compat) ─────────────────────────

@router.post("/generate-seating", status_code=status.HTTP_201_CREATED)
def generate_seating_route(
    payload: GenerateSeatingRequest,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Legacy endpoint: generate seating keyed to a single exam."""
    try:
        allocations = generate_seating(
            payload.exam_id,
            db,
            student_ids=payload.student_ids if payload.student_ids else None,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    halls_used = len({a.hall_id for a in allocations})
    return {
        "message": f"Seating generated. {len(allocations)} students allocated.",
        "total_allocated": len(allocations),
        "halls_used": halls_used,
    }


# ─── View seating by session ──────────────────────────────────────────────────

@router.get("/seating/session/{session_id}", response_model=SeatingResponse)
def get_session_seating(
    session_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Return all seat allocations for an exam session."""
    if not db.query(ExamSession).filter(ExamSession.id == session_id).first():
        raise HTTPException(status_code=404, detail="Exam session not found")

    rows = (
        db.query(SeatAllocation)
        .options(
            joinedload(SeatAllocation.student).joinedload(Student.department),
            joinedload(SeatAllocation.hall),
            joinedload(SeatAllocation.exam),
        )
        .filter(SeatAllocation.session_id == session_id)
        .order_by(SeatAllocation.hall_id, SeatAllocation.seat_number)
        .all()
    )

    details = [
        SeatAllocationDetail(
            id=row.id,
            seat_number=row.seat_number,
            hall_name=row.hall.name,
            student_name=row.student.full_name,
            register_number=row.student.register_number,
            department_name=row.student.department.name,
            exam_title=row.exam.title if row.exam else None,
        )
        for row in rows
    ]

    return SeatingResponse(
        session_id=session_id,
        total_allocated=len(details),
        allocations=details,
    )


# ─── Legacy view by exam_id ───────────────────────────────────────────────────

@router.get("/seating/{exam_id}", response_model=SeatingResponse)
def get_seating(
    exam_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Legacy: return seating allocations by exam_id."""
    if not db.query(Exam).filter(Exam.id == exam_id).first():
        raise HTTPException(status_code=404, detail="Exam not found")

    rows = (
        db.query(SeatAllocation)
        .options(
            joinedload(SeatAllocation.student).joinedload(Student.department),
            joinedload(SeatAllocation.hall),
            joinedload(SeatAllocation.exam),
        )
        .filter(SeatAllocation.exam_id == exam_id)
        .order_by(SeatAllocation.hall_id, SeatAllocation.seat_number)
        .all()
    )

    details = [
        SeatAllocationDetail(
            id=row.id,
            seat_number=row.seat_number,
            hall_name=row.hall.name,
            student_name=row.student.full_name,
            register_number=row.student.register_number,
            department_name=row.student.department.name,
            exam_title=row.exam.title if row.exam else None,
        )
        for row in rows
    ]

    return SeatingResponse(
        session_id=rows[0].session_id if rows else 0,
        total_allocated=len(details),
        allocations=details,
    )


# ─── Excel export by session ──────────────────────────────────────────────────

@router.get("/seating/session/{session_id}/export/excel")
def export_session_seating_excel(
    session_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Export seating for a session to Excel. One sheet per hall."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    session_obj = db.query(ExamSession).filter(ExamSession.id == session_id).first()
    if not session_obj:
        raise HTTPException(status_code=404, detail="Exam session not found")

    rows = (
        db.query(SeatAllocation)
        .options(
            joinedload(SeatAllocation.student).joinedload(Student.department),
            joinedload(SeatAllocation.hall),
            joinedload(SeatAllocation.exam),
        )
        .filter(SeatAllocation.session_id == session_id)
        .order_by(SeatAllocation.hall_id, SeatAllocation.seat_number)
        .all()
    )

    if not rows:
        raise HTTPException(status_code=404, detail="No seating allocations found for this session")

    # Group by hall
    halls_data = {}
    for row in rows:
        hall_name = row.hall.name
        if hall_name not in halls_data:
            halls_data[hall_name] = {
                "block": row.hall.block,
                "floor": row.hall.floor,
                "students": [],
            }
        halls_data[hall_name]["students"].append({
            "seat_number": row.seat_number,
            "register_number": row.student.register_number,
            "full_name": row.student.full_name,
            "department": row.student.department.name,
            "exam_title": row.exam.title if row.exam else "—",
        })

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for hall_name, data in halls_data.items():
        ws = wb.create_sheet(title=hall_name[:31])

        ws.merge_cells("A1:E1")
        ws["A1"] = f"Seating — {hall_name}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:E2")
        ws["A2"] = (
            f"Session: {session_obj.title}  |  "
            f"Date: {session_obj.exam_date}  |  "
            f"Time: {session_obj.start_time} – {session_obj.end_time}"
        )
        ws["A2"].alignment = Alignment(horizontal="center")

        hall_info = f"Hall: {hall_name}"
        if data["block"]:
            hall_info += f"  |  Block: {data['block']}"
        if data["floor"] is not None:
            hall_info += f"  |  Floor: {data['floor']}"
        ws.merge_cells("A3:E3")
        ws["A3"] = hall_info
        ws["A3"].alignment = Alignment(horizontal="center")

        headers = ["Seat No.", "Register No.", "Student Name", "Department", "Subject"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin

        for ri, student in enumerate(data["students"], 6):
            ws.cell(ri, 1, student["seat_number"]).border = thin
            ws.cell(ri, 2, student["register_number"]).border = thin
            ws.cell(ri, 3, student["full_name"]).border = thin
            ws.cell(ri, 4, student["department"]).border = thin
            ws.cell(ri, 5, student["exam_title"]).border = thin

        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 22
        ws.column_dimensions["E"].width = 30

    # Summary sheet
    summary = wb.create_sheet(title="Summary", index=0)
    summary["A1"] = "Exam Session Seating Summary"
    summary["A1"].font = Font(bold=True, size=14)
    summary["A3"] = "Session:"
    summary["B3"] = session_obj.title
    summary["A4"] = "Date:"
    summary["B4"] = str(session_obj.exam_date)
    summary["A5"] = "Time:"
    summary["B5"] = f"{session_obj.start_time} – {session_obj.end_time}"
    summary["A6"] = "Total Students:"
    summary["B6"] = len(rows)
    summary["A8"] = "Hall"
    summary["B8"] = "Students"
    summary["A8"].font = Font(bold=True)
    summary["B8"].font = Font(bold=True)
    for idx, (hall_name, data) in enumerate(halls_data.items(), 9):
        summary[f"A{idx}"] = hall_name
        summary[f"B{idx}"] = len(data["students"])
    summary.column_dimensions["A"].width = 30
    summary.column_dimensions["B"].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"seating_{session_obj.title.replace(' ', '_')}_{session_obj.exam_date}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── Legacy Excel export by exam ─────────────────────────────────────────────

@router.get("/seating/{exam_id}/export/excel")
def export_seating_excel(
    exam_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Legacy: export seating for a single exam."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(status_code=404, detail="Exam not found")

    rows = (
        db.query(SeatAllocation)
        .options(
            joinedload(SeatAllocation.student).joinedload(Student.department),
            joinedload(SeatAllocation.hall),
        )
        .filter(SeatAllocation.exam_id == exam_id)
        .order_by(SeatAllocation.hall_id, SeatAllocation.seat_number)
        .all()
    )

    if not rows:
        raise HTTPException(status_code=404, detail="No seating allocations found for this exam")

    halls_data = {}
    for row in rows:
        hall_name = row.hall.name
        if hall_name not in halls_data:
            halls_data[hall_name] = {"block": row.hall.block, "floor": row.hall.floor, "students": []}
        halls_data[hall_name]["students"].append({
            "seat_number": row.seat_number,
            "register_number": row.student.register_number,
            "full_name": row.student.full_name,
            "department": row.student.department.name,
        })

    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for hall_name, data in halls_data.items():
        ws = wb.create_sheet(title=hall_name[:31])
        ws.merge_cells("A1:D1")
        ws["A1"] = f"Seating Arrangement — {hall_name}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:D2")
        ws["A2"] = f"Exam: {exam.title}  |  Date: {exam.exam_date}  |  {exam.start_time}–{exam.end_time}"
        ws["A2"].alignment = Alignment(horizontal="center")
        for col, h in enumerate(["Seat No.", "Register No.", "Student Name", "Department"], 1):
            cell = ws.cell(5, col, h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin
        for ri, s in enumerate(data["students"], 6):
            ws.cell(ri, 1, s["seat_number"]).border = thin
            ws.cell(ri, 2, s["register_number"]).border = thin
            ws.cell(ri, 3, s["full_name"]).border = thin
            ws.cell(ri, 4, s["department"]).border = thin
        for col, w in zip("ABCD", [10, 18, 30, 25]):
            ws.column_dimensions[col].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"seating_{exam.title.replace(' ', '_')}_{exam.exam_date}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
