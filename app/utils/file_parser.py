"""
Utility functions for parsing CSV and Excel files for bulk upload.
"""
import csv
import io
from typing import List, Dict, Any

from fastapi import UploadFile, HTTPException


async def parse_upload_file(file: UploadFile) -> List[Dict[str, Any]]:
    """
    Parse an uploaded CSV or Excel file and return a list of row dictionaries.
    
    - Keys are normalized: lowercase, stripped of whitespace, underscores replace spaces.
    - Values are stripped of whitespace.
    - Empty rows are skipped.
    
    Raises HTTPException(400) if the file format is unsupported or parsing fails.
    """
    filename = file.filename.lower() if file.filename else ""
    
    if filename.endswith(".csv"):
        return await _parse_csv(file)
    elif filename.endswith(".xlsx"):
        return await _parse_xlsx(file)
    else:
        raise HTTPException(
            status_code=400,
            detail="Unsupported file format. Please upload a CSV (.csv) or Excel (.xlsx) file."
        )


async def _parse_csv(file: UploadFile) -> List[Dict[str, Any]]:
    """Parse a CSV file into a list of dictionaries."""
    try:
        content = await file.read()
        # Try to decode as UTF-8, fallback to latin-1
        try:
            text = content.decode("utf-8-sig")  # Handle BOM
        except UnicodeDecodeError:
            text = content.decode("latin-1")
        
        reader = csv.DictReader(io.StringIO(text))
        rows = []
        
        for row in reader:
            normalized = _normalize_row(row)
            if normalized:  # Skip empty rows
                rows.append(normalized)
        
        return rows
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Failed to parse CSV file: {str(e)}"
        )


async def _parse_xlsx(file: UploadFile) -> List[Dict[str, Any]]:
    """Parse an Excel (.xlsx) file into a list of dictionaries."""
    try:
        from openpyxl import load_workbook
        
        content = await file.read()
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        
        rows = []
        headers = []
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                # First row is headers
                headers = [_normalize_key(str(cell) if cell else "") for cell in row]
                continue
            
            # Build row dict
            row_dict = {}
            has_data = False
            for col_idx, cell in enumerate(row):
                if col_idx < len(headers) and headers[col_idx]:
                    value = str(cell).strip() if cell is not None else ""
                    row_dict[headers[col_idx]] = value
                    if value:
                        has_data = True
            
            if has_data:
                rows.append(row_dict)
        
        wb.close()
        return rows
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Failed to parse Excel file: {str(e)}"
        )


def _normalize_key(key: str) -> str:
    """Normalize a column header key: lowercase, strip, replace spaces with underscores."""
    return key.strip().lower().replace(" ", "_").replace("-", "_")


def _normalize_row(row: Dict[str, Any]) -> Dict[str, Any]:
    """Normalize all keys and values in a row dictionary."""
    normalized = {}
    for key, value in row.items():
        norm_key = _normalize_key(key)
        if norm_key:
            normalized[norm_key] = str(value).strip() if value else ""
    return normalized


def generate_csv_template(headers: List[str], example_rows: List[List[str]]) -> str:
    """
    Generate a CSV template string with headers and example rows.
    
    Args:
        headers: List of column header names
        example_rows: List of rows, each row is a list of values
    
    Returns:
        CSV string content
    """
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in example_rows:
        writer.writerow(row)
    return output.getvalue()
